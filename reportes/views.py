from django.http import HttpResponseRedirect
from django.views.generic import ListView, View, TemplateView
from django.urls import reverse_lazy
from django.contrib import messages
from django.shortcuts import render

#Librerias para genera el PDF
import os
from django.conf import settings
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from django.contrib.staticfiles import finders

#Libreria para generar el Excel
from openpyxl import Workbook

from parking.models import Factura, Empresa

# Create your views here.

class ListarFactura(ListView):
    # model = Factura
    # template_name = 'parking/listar_factura.html'
    # context_object_name = 'facturas'
    # queryset = Factura.objects.all().order_by('pk')

    def get(self, request, *args, **kwargs):
        queryset = Factura.objects.all().order_by('pk')
        context = {'facturas': queryset}
        return render(request, 'parking/listar_factura.html', context)

class GenerarPdfFactura(View):
    def get(self, request, *args, **kwargs):
        try:
            template = get_template('reportes/facturaUnica.html')
            context = {
                'factura': Factura.objects.get(pk = self.kwargs['pk']),
                'empresa': Empresa.objects.get(pk=1)
            }
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            # response['Content-Disposition'] = 'attachment; filename="report.pdf"'  Descargar automaticamente
            pisa_status = pisa.CreatePDF(html, dest=response)
            return response
        except Exception as e:
            messages.error(request, 'No se pudo generar el PDF, comunicarse con el administrador')
            return HttpResponseRedirect(reverse_lazy('parking:listar_factura'))

class GenerarInformeTotal(View):
    def get(self, request, *args, **kwargs):
        try:
            template = get_template('reportes/facturaConsolidada.html')
            context = {
                'facturas': Factura.objects.all(),
                'empresa': Empresa.objects.get(pk=1)
            }
            html = template.render(context)
            response = HttpResponse(content_type='application/pdf')
            pisa_status = pisa.CreatePDF(html, dest=response)
            return response
        except Exception as e:
            messages.error(request, 'No se pudo generar el PDF, comunicarse con el administrador')
            return HttpResponseRedirect(reverse_lazy('parking:listar_factura'))

class GenerarInformeTotalExcel(TemplateView):
    def get(self, request, *args, **kwargs):
        facturas = Factura.objects.all()
        wb = Workbook() #Creamos la instancia del Workbook
        ws = wb.active
        ws['A1'] = 'Reporte de Facturas'    #Casilla en la que queremos poner la informacion
        ws.merge_cells('A1:N1')

        ws['A2'] = 'numero factura'
        ws['B2'] = 'placa'
        ws['C2'] = 'categoria'
        ws['D2'] = 'tipo descuento'
        ws['E2'] = 'porcentaje descuento'
        ws['F2'] = 'tarifa'
        ws['G2'] = 'fecha ingreso'
        ws['H2'] = 'fecha salida'
        ws['I2'] = 'dias'
        ws['J2'] = 'horas'
        ws['K2'] = 'minutos'
        ws['L2'] = 'valor sin descuento'
        ws['M2'] = 'descuento'
        ws['N2'] = 'valor pagado'

        cont = 3                                                                    #Inicia el primer registro en la celda numero 3
        for factura in facturas:
            ws.cell(row = cont, column = 1).value = factura.pk                       #Row, son las filas , A,B,C,D osea row es igual al contador, y columnas 1,2,3
            ws.cell(row = cont, column = 2).value = factura.registroEntrada.placa.placa
            ws.cell(row = cont, column = 3).value = factura.registroEntrada.placa.tipo.tipo
            ws.cell(row = cont, column = 4).value = factura.registroEntrada.placa.descuento.tipoDescuento
            ws.cell(row = cont, column = 5).value = factura.registroEntrada.placa.descuento.porcentaje
            ws.cell(row = cont, column = 6).value = factura.registroEntrada.placa.tipo.tarifa
            ws.cell(row = cont, column = 7).value = factura.registroEntrada.horaIngreso
            ws.cell(row = cont, column = 8).value = factura.horaSalida
            ws.cell(row = cont, column = 9).value = factura.diasEstacionado
            ws.cell(row = cont, column = 10).value = factura.horasEstacionado
            ws.cell(row = cont, column = 11).value = factura.minutosEstaciondo
            ws.cell(row = cont, column = 12).value = factura.valorSinDescuento
            ws.cell(row = cont, column = 13).value = factura.descuento
            ws.cell(row = cont, column = 14).value = factura.valorPagar
            cont+=1
        nombre_archivo = "ReporteExcel.xlsx"
        response = HttpResponse(content_type = "application/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
